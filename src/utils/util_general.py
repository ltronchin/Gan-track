"""Miscellaneous utility classes and functions."""

import argparse
from pathlib import Path
import os
import numpy as np
import random
import openpyxl
import fnmatch
import torch
import requests
import collections
import sys
import shutil

from typing import Any, Union, List, Tuple

class iid_class:
    def __init__(self, iid_label=None):
        if iid_label is None:
            iid_label = [0, 2, 3, 4, 6, 7, 8, 9]
        self.iid_label = iid_label

    def idx_to_idx_iid_class(self, idx):
        assert idx <= len(self.iid_label)

        return self.iid_label[idx]

def get_idx_to_class(task_name, class_id):
    if task_name == "adaptive":
        return idx_to_class_ada(class_id)
    elif task_name == "overall_survival":
        return idx_to_class_os(class_id) # we provide in input classes of each network (ae_clinical, ae_images)
    else:
        raise ValueError(task_name)

def get_class_to_idx(task_name, class_name):
    if task_name == "adaptive":
        return class_to_idx_ada(class_name)
    elif task_name == "overall_survival":
        return class_to_idx_os(class_name) # we provide in input classes of each network (ae_clinical, ae_images)
    else:
        raise ValueError(task_name)

def idx_to_class_os(class_id):
    if class_id == 1:
        return 'death'
    elif class_id == 0:
        return 'alive'

def class_to_idx_os(class_name):
    if class_name == 'death':
        return 1
    elif class_name == 'alive':
        return 0

def idx_to_class_ada(class_id):
    if class_id == 1:
        return 'a'
    elif class_id == 0:
        return 'na'

def class_to_idx_ada(class_name):
    if class_name == 'a':
        return 1
    elif class_name == 'na':
        return 0

def list_dict():
   return collections.defaultdict(list)

def nested_dict():
   return collections.defaultdict(nested_dict)

def notification_ifttt(info):
    private_key = "isnY23hWBGyL-mF7F18BUAC-bGAN6dx1UAPoqnfntUa"
    url = "https://maker.ifttt.com/trigger/Notification/json/with/key/" + private_key
    requests.post(url, data={'value1': "Update --- "+ str(info)})

def define_source_path(path_dir, dataset, source_id_run=None, source_run_module=None):
    print('Define a source path')
    path_source_dir = path_dir

    print(f'Path parameters {path_dir}, {dataset}')
    if source_id_run is None:
        source_id_run = int(input("Enter the source   id_run  ."))
    assert type(source_id_run) == int
    if source_run_module is None:
        source_run_module = input("Enter the source   module  .")
    assert type(source_run_module) == str
    print(f'Path keys {source_id_run}, {source_run_module}')
    finded = False
    while not finded:
        source_run_name = "{0:05d}--{1}".format(source_id_run, source_run_module)
        path_source_dir = os.path.join(path_dir, dataset, source_run_name)
        if len(os.listdir(path_source_dir)) > 0:
            finded = True
            print('Source parameters in {}'.format(source_run_name))
            for datafile in os.listdir(path_source_dir):
                print(f"{datafile}")
        else:
            print(f'{source_id_run} or {source_run_module} not found! Try again')
            source_id_run = int(input("Enter the source   id_run  ."))
            source_run_module = input("Enter the source   module  .")
        print('Final source path: {}\n'.format(path_source_dir))
    return path_source_dir

def list_dir_recursively_with_ignore(dir_path: str, ignores: List[str] = None, add_base_to_relative: bool = False) -> List[Tuple[str, str]]:
    """List all files recursively in a given directory while ignoring given file and directory names.
    Returns list of tuples containing both absolute and relative paths."""
    assert os.path.isdir(dir_path)
    base_name = os.path.basename(os.path.normpath(dir_path))

    if ignores is None:
        ignores = []

    result = []

    for root, dirs, files in os.walk(dir_path, topdown=True):
        for ignore_ in ignores:
            dirs_to_remove = [d for d in dirs if fnmatch.fnmatch(d, ignore_)]

            # dirs need to be edited in-place
            for d in dirs_to_remove:
                dirs.remove(d)

            files = [f for f in files if not fnmatch.fnmatch(f, ignore_)]

        absolute_paths = [os.path.join(root, f) for f in files]
        relative_paths = [os.path.relpath(p, dir_path) for p in absolute_paths]

        if add_base_to_relative:
            relative_paths = [os.path.join(base_name, p) for p in relative_paths]

        assert len(absolute_paths) == len(relative_paths)
        result += zip(absolute_paths, relative_paths)

    return result

class Logger(object):
    """Redirect stderr to stdout, optionally print stdout to a file, and optionally force flushing on both stdout and the file."""

    def __init__(self, file_name: str = None, file_mode: str = "w", should_flush: bool = True):
        self.file = None

        if file_name is not None:
            self.file = open(file_name, file_mode)

        self.should_flush = should_flush
        self.stdout = sys.stdout
        self.stderr = sys.stderr

        sys.stdout = self
        sys.stderr = self

    def __enter__(self) -> "Logger":
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> None:
        self.close()

    def write(self, text: str) -> None:
        """Write text to stdout (and a file) and optionally flush."""
        if len(text) == 0: # workaround for a bug in VSCode debugger: sys.stdout.write(''); sys.stdout.flush() => crash
            return

        if self.file is not None:
            self.file.write(text)

        self.stdout.write(text)

        if self.should_flush:
            self.flush()

    def flush(self) -> None:
        """Flush written text to both stdout and a file, if open."""
        if self.file is not None:
            self.file.flush()

        self.stdout.flush()

    def close(self) -> None:
        """Flush, close possible files, and remove stdout/stderr mirroring."""
        self.flush()

        # if using multiple loggers, prevent closing in wrong order
        if sys.stdout is self:
            sys.stdout = self.stdout
        if sys.stderr is self:
            sys.stderr = self.stderr

        if self.file is not None:
            self.file.close()

def format_time(seconds: Union[int, float]) -> str:
    """Convert the seconds to human readable string with days, hours, minutes and seconds."""
    s = int(np.rint(seconds))

    if s < 60:
        return "{0}s".format(s)
    elif s < 60 * 60:
        return "{0}m {1:02}s".format(s // 60, s % 60)
    elif s < 24 * 60 * 60:
        return "{0}h {1:02}m {2:02}s".format(s // (60 * 60), (s // 60) % 60, s % 60)
    else:
        return "{0}d {1:02}h {2:02}m".format(s // (24 * 60 * 60), (s // (60 * 60)) % 24, (s // 60) % 60)


def copy_files_and_create_dirs(files: List[Tuple[str, str]]) -> None:
    """Takes in a list of tuples of (src, dst) paths and copies files.
    Will create all necessary directories."""
    for file in files:
        target_dir_name = os.path.dirname(file[1])

        # will create all intermediate-level directories
        if not os.path.exists(target_dir_name):
            os.makedirs(target_dir_name)

        shutil.copyfile(file[0], file[1])

def rgb2gray(rgb):
    return np.dot(rgb[..., :3], [0.2989, 0.5870, 0.1140])

def seed_all(seed): # for deterministic behaviour
    if not seed:
        seed = 42
    print("Using Seed : ", seed)

    os.environ['PYTHONHASHSEED'] = str(seed)
    torch.cuda.empty_cache()
    torch.manual_seed(seed) # Set torch pseudo-random generator at a fixed value
    torch.cuda.manual_seed_all(seed)
    torch.cuda.manual_seed(seed)
    np.random.seed(seed) # Set numpy pseudo-random generator at a fixed value
    random.seed(seed) # Set python built-in pseudo-random generator at a fixed value
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def create_dir(dir): # function to create directory
    if not os.path.exists(dir):
        Path(dir).mkdir(parents=True, exist_ok=True) # with parents 'True' creates all tree/nested folder

def create_path(*path_list, f=None):
    f = path_list[0]
    for i in range(1, len(path_list)):
        path = str(path_list[i])
        f = os.path.join(f, path)
    return f

def delete_file(file_path):
    try:
        os.remove(file_path)
    except FileNotFoundError:
        pass

def write_excel(*keys, dictionary, sheet_idx, col_position=1, wb=None, ws=None, default='not found'):
    """ Function to save the patients data splits (ids and labels) to file excel (openpyxl library is used)
    Args:
        keys: string values representing the field from the dictionary to save in excel
        dictionary: dictionary representing the data
        sheet_idx: int value representing the index of the current worksheet
        col_position: int value representing the coordinate of the column to begin to start the writing
        wb: existing workbook
        ws: existing worksheet
        default: default values to return if the inserted key do not exist in dictionary
    Returns:
        wb: workbook
        ws: worksheet
    """

    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '_fold' + str(sheet_idx)
    if ws is None:
        ws = wb.create_sheet(title='_fold' + str(sheet_idx))

    ws.cell(row=1, column=col_position, value=keys[0].split('_')[1] + '_id')
    ws.cell(row=1, column=col_position + 1, value=keys[0].split('_')[1] + '_label')

    for idx in range(dictionary.get(keys[0], default).shape[0]):
        [ws.cell(row=idx + 2, column=col_position + column, value=dictionary.get(key, default)[idx, 0]) for key, column in zip(keys, np.arange(len(keys)))]
    return wb, ws

def create_run_dir_local(run_dir_root) -> str:
    """Create a new run dir with increasing ID number at the start."""

    if not os.path.exists(run_dir_root):
        print("Creating the run dir root: {}".format(run_dir_root))
        os.makedirs(run_dir_root)

    run_id = get_next_run_id_local(run_dir_root)
    run_name = "{0:05d}".format(run_id)
    run_dir = os.path.join(run_dir_root, run_name)

    if os.path.exists(run_dir):
        raise RuntimeError("The run dir already exists! ({0})".format(run_dir))

    print("Creating the run dir: {}".format(run_dir))
    os.makedirs(run_dir)

    return run_dir


def get_next_run_id_local(run_dir_root: str, module_name: str) -> int:
    """Reads all directory names in a given directory (non-recursive) and returns the next (increasing) run id. Assumes IDs are numbers at the start of the directory names."""
    import re
    #dir_names = [d for d in os.listdir(run_dir_root) if os.path.isdir(os.path.join(run_dir_root, d))]
    #dir_names = [d for d in os.listdir(run_dir_root) if os.path.isdir(os.path.join(run_dir_root, d)) and d.split('--')[1] == module_name]
    dir_names = []
    for d in os.listdir(run_dir_root):
        if not 'configuration.yaml' in d and not 'log.txt' in d and not 'src' in d:
            if os.path.isdir(os.path.join(run_dir_root, d)) and d.split('--')[1] == module_name:
                dir_names.append(d)

    r = re.compile("^\\d+")  # match one or more digits at the start of the string
    run_id = 1

    for dir_name in dir_names:
        m = r.match(dir_name)

        if m is not None:
            i = int(m.group())
            run_id = max(run_id, i + 1)

    return run_id

def check_empty_directory(target_dir):
    pass